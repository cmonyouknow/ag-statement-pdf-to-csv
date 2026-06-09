import io
import re
import tempfile
from datetime import datetime
from pathlib import Path

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------

st.set_page_config(
	page_title="AG - Convert PDF Statement to Excel",
	page_icon="🔄",
	layout="centered",
)

st.title("AG - Convert PDF Statement to Excel")
st.caption("Allied Gold Ltd · PDF → XLSX")
st.markdown("Upload the statement from Iskra (old system) and convert to Excel. Please check before sending to client.")

# ---------------------------------------------------------------------------
# PDF parsing (identical logic to statement_to_xlsx.py)
# ---------------------------------------------------------------------------

def _parse_gbp(s: str) -> float:
	return float(s.replace("£", "").replace(",", "").replace(" ", "").strip())

def _parse_date_created(date_str: str) -> datetime:
	try:
		return datetime.strptime(date_str.strip(), "%d %b %Y")
	except ValueError:
		return datetime.today()

def parse_pdf(pdf_bytes: bytes) -> dict:
	with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
		tmp.write(pdf_bytes)
		tmp_path = tmp.name

	with pdfplumber.open(tmp_path) as pdf:
		all_lines = []
		for page in pdf.pages:
			text = page.extract_text()
			if text:
				all_lines.extend(text.split("\n"))

	data = {
		"title": "",
		"date_created": "",
		"from_name": "",
		"from_address": [],
		"to_name": "",
		"to_address": [],
		"section_label": "Manufacturing invoices",
		"invoices": [],
		"aging_labels": [],
		"aging_values": [],
		"overdue_amount": None,
		"total_amount": None,
	}

	for line in all_lines:
		if line.strip():
			data["title"] = line.strip()
			break

	for line in all_lines:
		m = re.search(r"Date created:\s+(.+)", line)
		if m:
			data["date_created"] = m.group(1).strip()
			break

	header_end = next((i for i, l in enumerate(all_lines) if "Invoice No." in l), 8)
	header_block = all_lines[1:header_end]

	from_parts, to_parts = [], []
	for line in header_block:
		if line.startswith("From:") or line.startswith("To:"):
			continue
		halves = re.split(r"\s{3,}", line.strip(), maxsplit=1)
		if len(halves) == 2:
			from_parts.append(halves[0].strip())
			to_parts.append(halves[1].strip())
		elif halves:
			to_parts.append(halves[0].strip())

	if from_parts:
		data["from_name"] = from_parts[0]
		data["from_address"] = from_parts[1:]
	if to_parts:
		data["to_name"] = to_parts[0]
		data["to_address"] = to_parts[1:]

	for line in all_lines:
		if line.strip().lower() in ("manufacturing invoices", "manufacturing invoice"):
			data["section_label"] = line.strip()
			break

	inv_pat = re.compile(
		r"^(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\S+)\s*(.*?)\s+"
		r"(£-?[\d,]+\.\d{2})\s+(£-?[\d,]+\.\d{2})$"
	)
	for line in all_lines:
		m = inv_pat.match(line.strip())
		if m:
			date, due, inv_no, desc, total_str, outstanding_str = m.groups()
			data["invoices"].append({
				"date": date, "due": due, "invoice_no": inv_no,
				"description": desc.strip(),
				"total": _parse_gbp(total_str),
				"outstanding": _parse_gbp(outstanding_str),
			})

	for i, line in enumerate(all_lines):
		if "0-30 Days" in line or "0–30 Days" in line:
			data["aging_labels"] = re.split(r"\s{2,}", line.strip())
			if i + 1 < len(all_lines):
				raw_vals = re.findall(r"£\s*-?[\d,]+\.\d{2}", all_lines[i + 1])
				data["aging_values"] = [_parse_gbp(v.replace(" ", "")) for v in raw_vals]
			break

	for line in all_lines:
		m = re.search(r"Overdue Amount:\s*(£\s*-?[\d,]+\.\d{2})", line)
		if m:
			data["overdue_amount"] = _parse_gbp(m.group(1))
		m = re.search(r"Total Amount:\s*(£\s*-?[\d,]+\.\d{2})", line)
		if m:
			data["total_amount"] = _parse_gbp(m.group(1))

	return data

# ---------------------------------------------------------------------------
# XLSX generation (identical logic to statement_to_xlsx.py)
# ---------------------------------------------------------------------------

LIGHT_GRAY = "E8E8E8"
ORANGE_RED = "CC3300"
RED = "CC0000"
BLACK = "000000"
GBP_FMT = "£#,##0.00"

def _nb(): return Side(style=None)
def _side(style="thin", color="999999"): return Side(style=style, color=color)
def no_border(): return Border(left=_nb(), right=_nb(), top=_nb(), bottom=_nb())
def bottom_only(color="CCCCCC"): return Border(left=_nb(), right=_nb(), top=_nb(), bottom=_side("dotted", color))
def top_bottom_border():
	s = _side("thin", BLACK)
	return Border(left=_nb(), right=_nb(), top=s, bottom=s)
def all_thin_border(color="AAAAAA"):
	s = _side("thin", color)
	return Border(left=s, right=s, top=s, bottom=s)
def top_only_border(): return Border(left=_nb(), right=_nb(), top=_side("thin", "999999"), bottom=_nb())

def build_xlsx(data: dict) -> bytes:
	wb = Workbook()
	ws = wb.active
	ws.title = "Statement"
	ws.sheet_view.showGridLines = False

	for col, width in zip("ABCDEF", [13, 13, 18, 18, 16, 16]):
		ws.column_dimensions[col].width = width

	row = 1

	ws.merge_cells(f"A{row}:F{row}")
	c = ws[f"A{row}"]
	c.value = data["title"]
	c.font = Font(name="Arial", bold=True, size=14)
	c.alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[row].height = 24
	row += 1

	ws.row_dimensions[row].height = 6
	row += 1

	for label, col_letter in [("From:", "A"), ("To:", "C")]:
		c = ws[f"{col_letter}{row}"]
		c.value = label
		c.font = Font(name="Arial", bold=True, size=10)
	ws[f"E{row}"] = "Date created:"
	ws[f"E{row}"].font = Font(name="Arial", bold=True, size=10)
	ws[f"F{row}"] = data["date_created"]
	ws[f"F{row}"].font = Font(name="Arial", bold=True, size=10)
	ws.row_dimensions[row].height = 15
	row += 1

	from_lines = [data["from_name"]] + data["from_address"]
	to_lines = [data["to_name"]] + data["to_address"]
	for i in range(max(len(from_lines), len(to_lines))):
		fl = from_lines[i] if i < len(from_lines) else ""
		tl = to_lines[i] if i < len(to_lines) else ""
		ws[f"A{row}"] = fl
		ws[f"A{row}"].font = Font(name="Arial", size=10, bold=(i == 0))
		ws[f"C{row}"] = tl
		ws[f"C{row}"].font = Font(name="Arial", size=10, bold=(i == 0))
		ws.row_dimensions[row].height = 14
		row += 1

	ws.row_dimensions[row].height = 6
	row += 1

	for col, header in enumerate(["Date", "Due", "Invoice No.", "Description", "Total", "Outstanding"], 1):
		c = ws.cell(row=row, column=col, value=header)
		c.font = Font(name="Arial", bold=True, size=10)
		c.fill = PatternFill("solid", start_color=LIGHT_GRAY)
		c.alignment = Alignment(horizontal="left" if col <= 4 else "right", vertical="center")
		c.border = no_border()
	ws.row_dimensions[row].height = 18
	row += 1

	ws.merge_cells(f"A{row}:F{row}")
	c = ws[f"A{row}"]
	c.value = data["section_label"]
	c.font = Font(name="Arial", bold=True, size=10)
	c.alignment = Alignment(horizontal="left")
	c.border = bottom_only("CCCCCC")
	ws.row_dimensions[row].height = 16
	row += 1

	data_start = row
	stmt_date = _parse_date_created(data["date_created"])

	for inv in data["invoices"]:
		due_dt = datetime.strptime(inv["due"], "%d/%m/%Y")
		is_overdue = due_dt <= stmt_date
		vals = [inv["date"], inv["due"], inv["invoice_no"], inv["description"], inv["total"], inv["outstanding"]]
		for col, val in enumerate(vals, 1):
			c = ws.cell(row=row, column=col, value=val)
			c.border = bottom_only()
			c.font = Font(name="Arial", size=10)
			if col == 2:
				c.font = Font(name="Arial", size=10, bold=is_overdue, color=ORANGE_RED if is_overdue else BLACK)
			if col in (5, 6):
				c.number_format = GBP_FMT
				c.alignment = Alignment(horizontal="right")
			else:
				c.alignment = Alignment(horizontal="left")
		ws.row_dimensions[row].height = 15
		row += 1

	ws.cell(row=row, column=1, value="Subtotal").font = Font(name="Arial", bold=True, size=10)
	ws.cell(row=row, column=1).border = top_bottom_border()
	for col in range(2, 5):
		ws.cell(row=row, column=col).border = top_bottom_border()
	for col in (5, 6):
		cl = get_column_letter(col)
		c = ws.cell(row=row, column=col)
		c.value = f"=SUM({cl}{data_start}:{cl}{row - 1})"
		c.font = Font(name="Arial", bold=True, size=10)
		c.number_format = GBP_FMT
		c.alignment = Alignment(horizontal="right")
		c.border = top_bottom_border()
	ws.row_dimensions[row].height = 18
	row += 1

	ws.row_dimensions[row].height = 8
	row += 1

	if data["aging_labels"]:
		for col, lbl in enumerate(data["aging_labels"], 1):
			c = ws.cell(row=row, column=col, value=lbl)
			c.font = Font(name="Arial", bold=True, size=10)
			c.fill = PatternFill("solid", start_color=LIGHT_GRAY)
			c.alignment = Alignment(horizontal="left")
			c.border = all_thin_border()
		for col in range(len(data["aging_labels"]) + 1, 7):
			ws.cell(row=row, column=col).fill = PatternFill("solid", start_color=LIGHT_GRAY)
			ws.cell(row=row, column=col).border = all_thin_border()
		ws.row_dimensions[row].height = 16
		row += 1

		for col, val in enumerate(data["aging_values"], 1):
			c = ws.cell(row=row, column=col, value=val)
			c.font = Font(name="Arial", size=10)
			c.number_format = GBP_FMT
			c.alignment = Alignment(horizontal="left")
			c.border = all_thin_border()
		for col in range(len(data["aging_values"]) + 1, 7):
			ws.cell(row=row, column=col).border = all_thin_border()
		ws.row_dimensions[row].height = 16
		row += 1

	ws.row_dimensions[row].height = 8
	row += 1

	if data["overdue_amount"] is not None:
		ws.cell(row=row, column=5, value="Overdue Amount:").font = Font(name="Arial", bold=True, size=10, color=RED)
		ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
		c = ws.cell(row=row, column=6, value=data["overdue_amount"])
		c.font = Font(name="Arial", bold=True, size=10, color=RED)
		c.number_format = GBP_FMT
		c.alignment = Alignment(horizontal="right")
		ws.row_dimensions[row].height = 16
		row += 1

	if data["total_amount"] is not None:
		ws.cell(row=row, column=5, value="Total Amount:").font = Font(name="Arial", bold=True, size=10)
		ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
		c = ws.cell(row=row, column=6, value=data["total_amount"])
		c.font = Font(name="Arial", bold=True, size=10)
		c.number_format = GBP_FMT
		c.alignment = Alignment(horizontal="right")
		ws.row_dimensions[row].height = 16
		row += 1

	ws.row_dimensions[row].height = 8
	row += 1

	for i, line in enumerate([
		"Allied Gold Ltd, Elma House, Beaconsfield Close, Hatfield AL10 8YG, Tel.: 020 7249 4422, Fax: 020 7249 1166, VAT Number: 577546200",
		"Bank Details: Barclays Bank Plc, Sort code: 20-12-83, Account Number: 43420159, SWIFT: BARCGB22, IBAN: GB91BARC20128343420159",
	]):
		ws.merge_cells(f"A{row}:F{row}")
		c = ws[f"A{row}"]
		c.value = line
		c.font = Font(name="Arial", bold=(i == 1), size=8, color="555555")
		c.alignment = Alignment(horizontal="center")
		if i == 0:
			c.border = top_only_border()
		ws.row_dimensions[row].height = 14
		row += 1

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()

# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

uploaded = st.file_uploader("Drop your PDF statement here", type="pdf")

if uploaded:
	with st.spinner("Reading PDF..."):
		try:
			data = parse_pdf(uploaded.read())
		except Exception as e:
			st.error(f"Could not parse the PDF: {e}")
			st.stop()

	if not data["invoices"]:
		st.error("No invoices found — is this an Allied Gold statement?")
		st.stop()

	# Summary metrics
	st.success(f"Found **{len(data['invoices'])} invoices** in the statement.")
	col1, col2, col3 = st.columns(3)
	col1.metric("Total", f"£{data['total_amount']:,.2f}" if data["total_amount"] else "—")
	col2.metric("Overdue", f"£{data['overdue_amount']:,.2f}" if data["overdue_amount"] else "—")
	col3.metric("Date created", data["date_created"] or "—")

	with st.spinner("Building XLSX..."):
		xlsx_bytes = build_xlsx(data)

	stem = Path(uploaded.name).stem
	out_name = f"{stem}.xlsx"

	st.download_button(
		label="⬇️ Download Excel file",
		data=xlsx_bytes,
		file_name=out_name,
		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		use_container_width=True,
		type="primary",
	)